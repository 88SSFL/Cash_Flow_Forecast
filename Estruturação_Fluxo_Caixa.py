from Suporte_Fluxo_de_caixa_teste import *
import pandas as pd

#endereço= r'C:\Users\lippe\Downloads\re\2022.12.05 - DIVR007 - Planilha Indicadores.xlsx'
endereço=r'C:\Users\u3xp\OneDrive - PETROBRAS\Desktop\2023.01.16 - DIVR007 - Planilha Indicadores.xlsx'
#%%
cdi_dataframe = pd.read_excel(endereço, 
                              sheet_name='Indicadores', usecols=[0,1 ,3,4,5,6,7,10,16,21,30,34,36,37])

col = cdi_dataframe.columns

cdi_dataframe.loc[:,col[0]] = st_dt(cdi_dataframe.loc[:,col[0]])
b_day_=b_day(cdi_dataframe)

#%%
amortizacao_teste = pd.DataFrame(
    {'Data': ['15/08/2023'], 'Valores': [176969062.345892]})
#%%
def cash_flow(type_, nome, moeda, data_base, emissão, maturity, periodo_cupom, valor_emissão,  cupom,
              indexador, amortização=0, tipo_de_indexação=0 ):
    if amortização.empty==True:
        amortização= pd.DataFrame({'Data':[ f'{emissão}'], 'Valores': ['0']})
    data_base, emissão, maturity = dt_(data_base), dt_(emissão), dt_(maturity)
    freq = 12 / periodo_cupom
    amortização.iloc[:, 0] = st_dt(amortização.iloc[:, 0])
    cl = creating_lists(data_base, maturity, emissão, freq, amortização, valor_emissão,indexador,b_day_)
    dates, amortization, outstanding, coupon_dates = cl[0], cl[1], cl[2], cl[3]
    interest_treatment=_interest_(type_, amortization, outstanding, emissão, coupon_dates, dates, freq, data_base, cupom, tipo_de_indexação, indexador, nome)
    coupon_values,outstanding, amortization = interest_treatment[0],interest_treatment[1],interest_treatment[2]
    if type_ == 'ipca': ipca=interest_treatment[3]
    name = [nome] * len(dates)
    cash_flow = conclusion_origin_currency(dates,emissão, name, amortization, outstanding, coupon_values, indexador,type_)
    if type_ in ['ipca','cdi','tjlp','fixo_360','fixo_365','fixo_util','mapa4']:
        moeda='BRL'
    if moeda == 'BRL':
        swap_ = swap(indexador, cash_flow,type_,coupon_dates,data_base,dates,freq, nome, "no")
        cash_flow_usd=swap_[0]
    else:
        swap_=[fx(moeda, indexador, cash_flow,data_base,coupon_dates,dates), "empty"]
        cash_flow_usd = swap_[0]
    cash_flow = conclusion(cash_flow, cash_flow_usd,indexador)
    return [cash_flow,swap_[1]]

 #%%

bpc=pd.read_excel(r'C:\Users\u3xp\OneDrive - PETROBRAS\Documents\2023.01.16 - DIVR006 - Fluxo de Caixa_new.xlsx'
                  ,usecols=[1,10,21,22,23,24,25,26,27,28,29], sheet_name='Fluxo de Caixa')
#%%
data_base_debt= pd.read_excel(r'C:\Users\u3xp\OneDrive - PETROBRAS\Documents\teste_fluxo_caixa.xlsx',
                              sheet_name='Planilha1')
data_base_amortization=pd.read_excel(r'C:\Users\u3xp\OneDrive - PETROBRAS\Documents\teste_fluxo_caixa.xlsx',
                              sheet_name='Planilha2')

#'libor_6','libor_3','sofr_6','sofr_3'
#data_base_debt=data_base_debt[data_base_debt.iloc[:,0]=='ipca']

initial_database_date=cdi_dataframe.iloc[0,0]
#%%
from openpyxl import load_workbook,workbook
from openpyxl.utils import get_column_letter
import time
t1=time.time()
to_sheet=list()
len_=len(data_base_debt)
data_base="16/01/2023"

for i in range(len_):
    db=data_base_debt.iloc[i,:]
    if all([dt_(data_base)>=dt_(str(db.iloc[3])[:10]),dt_(data_base)<dt_(str(db.iloc[4])[:10])]):
        print(db.iloc[1].upper())
        amort=data_base_amortization[data_base_amortization.loc[:,'nome']==db.iloc[1]].iloc[:,1:].reset_index(drop=True)
        amort.iloc[:,0]=list(map(lambda x:str(x)[:10],amort.iloc[:,0]))
        amort.iloc[:,1]=list(map(lambda x:-float(x),amort.iloc[:,1]))
        db.iloc[3]=str(db.iloc[3])
    
        if dt_(db.iloc[3])<initial_database_date: db.iloc[3]=datetime(2018,dt_(db.iloc[3]).month,dt_(db.iloc[3]).day)
        intermediary=cash_flow(db.iloc[0], db.iloc[1], db.iloc[2], data_base, str(db.iloc[3])[:10], str(db.iloc[4])[:10], db.iloc[5], db.iloc[6], db.iloc[7], cdi_dataframe, amort, db.iloc[8])
        intermediary_cf=intermediary[0]
        final=intermediary_cf if len(to_sheet)==0 else pd.concat([final,intermediary_cf], ignore_index=True)
        to_sheet.append([intermediary_cf,db.iloc[1],bpc[bpc.loc[:,'Portfólio']==db.iloc[1]],intermediary[1]])
        
#%%
try:
    wb_0=load_workbook(r'C:\Users\u3xp\OneDrive - PETROBRAS\Documents\cash flow.xlsx')
except:
    to_sheet[0][0].to_excel(r'C:\Users\u3xp\OneDrive - PETROBRAS\Documents\cash flow.xlsx')
    wb_0=load_workbook(r'C:\Users\u3xp\OneDrive - PETROBRAS\Documents\cash flow.xlsx')
    del wb_0[wb_0.sheetnames[0]] 
with pd.ExcelWriter(r'C:\Users\u3xp\OneDrive - PETROBRAS\Documents\cash flow.xlsx', engine = 'openpyxl') as writer:
    writer.book=wb_0
    for i in to_sheet:
        i[0].to_excel(writer,sheet_name=i[1],startcol=0, index=False)
        i[2].to_excel(writer,sheet_name=i[1],startcol=15, index=False)
        ws = wb_0[wb_0.sheetnames[-1]]
        for l in range(len(i[0].columns)):
            if l==0:
                for k in range(len(i[0])):
                    ws[get_column_letter(l+1)+str(k+2)].number_format='dd/mm/yyyy'
                ws.column_dimensions['A'].width=12
            elif l==1:
                max_=0
                for k in range(len(i[0])):
                    max_=max(max_,len(i[0].iloc[k,1]))
                ws.column_dimensions['B'].width=max_+3
            elif l in range(2,11):
                for k in range(len(i[0])):
                    ws[get_column_letter(l+1)+str(k+2)].value=float(ws[get_column_letter(l+1)+str(k+2)].value)
                    ws[get_column_letter(l+1)+str(k+2)].number_format='#,##0.00'
                ws.column_dimensions[get_column_letter(l+1)].width=max(list(map(lambda x: len(str(x*100//1/100)),i[0].iloc[:,l])))+5
            else:
                break
        for m in range(len(i[2].columns)):
            if m==0:
                 max_=0
                 for n in range(len(i[2])):
                    max_=max(max_,len(i[2].iloc[n,0]))
                 ws.column_dimensions['P'].width=max_+3               
            elif m==1:               
                for n in range(len(i[2])):
                    ws[get_column_letter(m+16)+str(n+2)].number_format='dd/mm/yyyy'
                ws.column_dimensions['Q'].width=12
            elif m in range(2,11):
                for n in range(len(i[2])):
                    ws[get_column_letter(m+16)+str(n+2)].value=float(ws[get_column_letter(m+16)+str(n+2)].value) if  ws[get_column_letter(m+16)+str(n+2)].value!="" else 0.00
                    ws[get_column_letter(m+16)+str(n+2)].number_format='#,##0.00'
                ws.column_dimensions[get_column_letter(m+16)].width=max(list(map(lambda x: len(str(x*100//1/100)),i[2].iloc[:,m])))+5
            else:
                break
        for j in range(9):
            for k in range(max(len(i[0]),len(i[2]))):
                if all([ws[get_column_letter(j+3)+str(k+2)].value ==0,ws[get_column_letter(j+18)+str(k+2)].value ==0]):
                    ws[get_column_letter(j+27)+str(k+2)].value= "-"
                elif all([ws[get_column_letter(j+3)+str(k+2)].value not in [0, None], ws[get_column_letter(j+18)+str(k+2)].value not in [0, None]]):
                    if all([ws[get_column_letter(j+18)+str(k+2)].value/ws[get_column_letter(j+3)+str(k+2)].value>0.95,
                         ws[get_column_letter(j+18)+str(k+2)].value/ws[get_column_letter(j+3)+str(k+2)].value<1.05]):
                        ws[get_column_letter(j+27)+str(k+2)].value= "-"  
                    else:
                         ws[get_column_letter(j+27)+str(k+2)].value="erro"
                else:
                    ws[get_column_letter(j+27)+str(k+2)].value="erro"
        ws.sheet_view.zoomScale = 85
            
#%%
from openpyxl import load_workbook,workbook
final.to_excel(r'C:\Users\u3xp\OneDrive - PETROBRAS\Documents\cash flow unico.xlsx', index=False)
interest_ = 0
for i in range(1, 14):    
    while True:
        interest_ += 10 ** (-i + 1)
        if sum(list(map(lambda x,y:x/((1+interest_/100)**(y-final.iloc[0,0]).days),final.iloc[:,9],final.iloc[:,0])))>0:
           interest_ -= 10 ** (-i + 1)
           break
interest=((1+interest_/100)**365-1)*100
print(interest)
avglife=(sum(final.iloc[:,3]*(list(map(lambda x:(x-final.iloc[0,0]).days,final.iloc[:,0]))))/sum(final.iloc[:,3]))/365
wb_=load_workbook(r'C:\Users\u3xp\OneDrive - PETROBRAS\Documents\cash flow unico.xlsx')
ws_1 = wb_[wb_.sheetnames[-1]]
ws_1['P2'].value,ws_1['Q2'].value='Prazo Médio (anos)', float(avglife)
ws_1['P3'].value,ws_1['Q3'].value='Custo Médio (%)', float(interest)

ws_1.column_dimensions['P'].width=19
ws_1.column_dimensions['Q'].width=10
ws_1['Q2'].number_format="0.000"
ws_1['Q3'].number_format="0.000"

for i in range(1,12):
    if i==1:
        for j in range(len(final)):
            ws_1['A'+str(j+2)].number_format='dd/mm/yyyy'
        ws_1.column_dimensions['A'].width=12
    elif i==2:
        max_=0
        for j in range(len(final)):
            max_=max(max_,len(final.iloc[j,1]))
        ws_1.column_dimensions['B'].width=max_+3
    else:
        width_3=len(list(str(max(list(map(lambda x: -x//1 if x<0 else x//1,final.iloc[:,i-1]))))))+5
        ws_1.column_dimensions[get_column_letter(i)].width=width_3
        for j in range(2,len(final)+2):
            ws_1[get_column_letter(i)+str(j)].value=float(ws_1[get_column_letter(i)+str(j)].value)
            ws_1[get_column_letter(i)+str(j)].number_format='#,##0.00'


wb_.save(r'C:\Users\u3xp\OneDrive - PETROBRAS\Documents\cash flow unico.xlsx')
wb_.close()        
#%%
swap_0,swap_1,swap_2=[],[],[]
for i in range(len(to_sheet)):
    if to_sheet[i][-1]!="empty":
        swap_bpc=(((to_sheet[i][2].iloc[-1,-4]/to_sheet[i][2].iloc[-1,-8]+1)
                   **(1/(dt_(to_sheet[i][2].iloc[-1,1])-dt_(to_sheet[i][2].iloc[-2,1])).days)-1)     *36000)
        swap_0.append(to_sheet[i][1])
        swap_1.append(to_sheet[i][-1])
        swap_2.append(swap_bpc)

swap = pd.DataFrame({'Nome':swap_0, 'Swap_Python':swap_1,'Swap_BPC':swap_2})
swap.to_excel(r'C:\Users\u3xp\OneDrive - PETROBRAS\Documents\comparativo swap.xlsx')
#%%
npv_0,npv_1,npv_2, npv_3,npv_4=[],[],[],[],[]
for i in range(len(to_sheet)):
    npv_0.append(to_sheet[i][1])
    npv_1.append(sum(to_sheet[i][0].iloc[:,10]))
    npv_2.append(sum(to_sheet[i][2].iloc[:,10]))
    npv_3.append(npv_1[-1]/to_sheet[i][0].iloc[0,10])   
    npv_4.append(npv_2[-1]/to_sheet[i][2].iloc[0,10] if len(to_sheet[i][2])>0 else 0)
swap = pd.DataFrame({'Nome':npv_0, 'VPL_Python':npv_1,'VPL_BPC':npv_2, 'VPL/principal Python': npv_3, 'VPL/principal BPC': npv_4})
swap.to_excel(r'C:\Users\u3xp\OneDrive - PETROBRAS\Documents\comparativo vpl.xlsx')
wb_3=load_workbook(r'C:\Users\u3xp\OneDrive - PETROBRAS\Documents\comparativo vpl.xlsx')
ws_3 = wb_3[wb_3.sheetnames[-1]]

for i in range(3,7):
    for j in range(len(npv_0)):
        if i in [3,4]:
             ws_3[get_column_letter(i)+str(j+2)].value=float(ws_3[get_column_letter(i)+str(j+2)].value)
             ws_3[get_column_letter(i)+str(j+2)].number_format='#,##0.00'
        else:
            ws_3[get_column_letter(i)+str(j+2)].value=float(ws_3[get_column_letter(i)+str(j+2)].value)
            ws_3[get_column_letter(i)+str(j+2)].number_format='0.000%'
             
ws_3.column_dimensions['B'].width=29
ws_3.column_dimensions['C'].width=len(str(max(npv_1)))
ws_3.column_dimensions['D'].width=len(str(max(npv_2)))
ws_3.column_dimensions['E'].width=20 
ws_3.column_dimensions['F'].width=20        
            
wb_3.save(r'C:\Users\u3xp\OneDrive - PETROBRAS\Documents\comparativo vpl.xlsx')
wb_.close()


t2=time.time()        
print(t2-t1)