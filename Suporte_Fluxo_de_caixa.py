import pandas as pd
import numpy as np
from datetime import datetime
from dateutil.relativedelta import relativedelta

# Indexing treatment defs

def ipca_indexation(indexador, outstanding,amortization,dates, data_base):
    col = list(indexador.columns)
    suport_date = list(indexador[indexador.loc[:,col[1]] == 1].iloc[0, [0,3]])
    ipca_ = indexador[indexador.loc[:,col[0]].isin(dates)].loc[:,[col[0], col[3], 'CURVA IPCA.1']]
    ipca_.loc[:,'index'] = list(   
        map(lambda x, y: round(round((1 + x) ** (252 / 22),7)**(22/252),7)**((y - suport_date[1]) / 22) 
            , ipca_.iloc[:, 2], ipca_.iloc[:, 1]))          
    for i in range((1 if suport_date[0] ==data_base else 0),len(outstanding)):
        outstanding[i]=outstanding[i]* ipca_.iloc[i, -1]
        amortization[i]=amortization[i]*ipca_.iloc[i, -1]
    amortization[-1]=-outstanding[-2]* ipca_.iloc[-1, -1]/ipca_.iloc[-2, -1]
    return [outstanding,amortization,ipca_]

def tjlp_indexation(indexador,coupon_dates,dates,outstanding,amortization, data_base):
    col = list(indexador.columns)
    suport_date = indexador[indexador.loc[:,col[1]] == 1].iloc[0, 0]
    ind = indexador[indexador.loc[:,'PERÍODO'].isin(coupon_dates)].loc[:,['PERÍODO', 'TJLP(%).1']]
    tjlp_=[1]+list(map(lambda x,y,z:((1+x/100)/1.06)**((y-z).days/360) if x>6 else 1,ind.iloc[:,1], ind.iloc[1:,0], ind.iloc[:-1,0]))
    control_3_1, control_3_2=0,0
    while True:
        if coupon_dates[control_3_1]<dates[control_3_2]:
            control_3_1+=1
        elif coupon_dates[control_3_1]==dates[control_3_2]:
            control_3_1+=1
            control_3_2+=1
        else:
            tjlp_.insert(control_3_1,(((1+ind.iloc[control_3_1-1,1]/100)/1.06) **((dates[control_3_2]-coupon_dates[control_3_1-1]).days/360) if ind.iloc[control_3_1-1,1]>6 else 1 ))
            control_3_2+=1
        if control_3_1==len(coupon_dates) or control_3_2==len(dates):
            break
    tjlp_=np.asarray(tjlp_)
    for i in range((1 if suport_date ==data_base else 0),len(outstanding)):
        amortization[i]=amortization[i]/(outstanding[i]-amortization[i])
        outstanding[i]=outstanding[i-1]*tjlp_[i]*(1+amortization[i])
        amortization[i]=outstanding[i]*amortization[i]/(1+amortization[i])
    amortization[-1]=-outstanding[-2]*tjlp_[-1]
    outstanding[-1]=0
    return [outstanding,amortization]

def mapa4_indexation (coupon, freq, coupon_dates, dates, outstanding, amortization):
    coupon_indexation=[(1+coupon/100)**((12-freq)/12)]*(len(coupon_dates)-1)
    control_3_1, control_3_2=0,0
    while True:
        if coupon_dates[control_3_1]<dates[control_3_2]:
            control_3_1+=1
        elif coupon_dates[control_3_1]==dates[control_3_2]:
            control_3_1+=1
            control_3_2+=1
        else:
            coupon_indexation.insert(control_3_1-1,coupon_indexation[control_3_1-1]**((dates[control_3_2]-coupon_dates[control_3_1-1]).days/
                                                                     (coupon_dates[control_3_1]-coupon_dates[control_3_1-1]).days))
            coupon_indexation[control_3_1]=coupon_indexation[control_3_1]/coupon_indexation[control_3_1-1]
            control_3_2+=1
        if control_3_1==len(coupon_dates) or control_3_2==len(dates):
            break
    coupon_indexation=coupon_indexation[-len(outstanding):]
    for i in range(1,len(outstanding)):
        amortization[i]=amortization[i]/(outstanding[i]-amortization[i])
        outstanding[i]=outstanding[i-1]*coupon_indexation[i]*(1+amortization[i])
        amortization[i]=outstanding[i]*amortization[i]/(1+amortization[i])
        
    amortization[-1]=-outstanding[-2]*coupon_indexation[-1]
    outstanding[-1]=0
    return[outstanding, amortization]
    
# COUPON VALUES defs

def pv(indexador,data_base,coupon_dates,dates,cashflow):
    ind=indexador
    next_date_coupon=list(filter(lambda x: x>data_base,coupon_dates))[0]
    last_date_coupon=list(filter(lambda x: x<=data_base,coupon_dates))[-1]
    period_interest=1-(cashflow.iloc[dates.index(next_date_coupon),4]/(cashflow.iloc[dates.index(next_date_coupon),3]-cashflow.iloc[dates.index(next_date_coupon),2]))
    days_gone=ind.iloc[index(ind, data_base, 0), 3]- ind.iloc[index(ind, last_date_coupon, 0), 3]
    days_total=ind.iloc[index(ind, next_date_coupon, 0), 3]-ind.iloc[index(ind, last_date_coupon,0),3]
    return cashflow.iloc[dates.index(data_base),5]*period_interest**(days_gone/days_total)
    
def bond_coupon(amortization, outstanding, control_2, cupom, freq, dates, coupon_dates):
    principal = (amortization[1:] - outstanding[1:] * control_2[-len(amortization)+1:])
    coupon_interest = ((cupom / (freq * 100)) + 1)  
    time = list(map(lambda x, y: ((coupon_interest ** (((x - y).days / (360 / freq)))) - 1), coupon_dates[1:], coupon_dates[:-1]))
    auxiliary_=dates[0]
    for i in range(1,len(dates)):
        if control_2[i]==True:
            auxiliary_=dates[i]
        else:
            time.insert(i-1,((coupon_interest ** (((dates[i] - auxiliary_).days / (360 / freq)))) - 1))
    return [0]+(principal * time[-len(dates)+1:]).tolist()

def fixo_360_coupon(amortization, outstanding, control_2, cupom,dates, coupon_dates):
    principal = (amortization[1:] - outstanding[1:] * control_2[-len(amortization)+1:])
    time = list(map(lambda x, y: (cupom/100)* ((x - y).days / 360 ) , coupon_dates[1:], coupon_dates[:-1]))
    auxiliary_=dates[0]
    for i in range(1,len(dates)):
        if control_2[i]==True:
            auxiliary_=dates[i]
        else:
            time.insert(i-1,(cupom/100) * ((dates[i] - auxiliary_).days / 360 ))
    return [0]+(principal * time[-len(dates)+1:]).tolist()

def fixo_365_coupon(amortization, outstanding, control_2, cupom,dates, coupon_dates):
    principal = (amortization[1:] - outstanding[1:] * control_2[-len(amortization)+1:])
    time = list(map(lambda x, y: (cupom/100)* ((x - y).days / 365 ) , coupon_dates[1:], coupon_dates[:-1]))
    auxiliary_=dates[0]
    for i in range(1,len(dates)):
        if control_2[i]==True:
            auxiliary_=dates[i]
        else:
            time.insert(i-1,(cupom/100) * ((dates[i] - auxiliary_).days / 365 ))
    return [0]+(principal * time[-len(dates)+1:]).tolist()

def fixo_util_coupon(amortization, outstanding, control_2, cupom,dates, coupon_dates,indexador):
    ind = indexador
    principal = (amortization[1:] - outstanding[1:] * control_2[-len(amortization)+1:])
    time = list(map(lambda x, y: ((1+cupom/100)**((ind.iloc[index(ind, x, 0), 3]- ind.iloc[index(ind, y, 0), 3])/ 252 )-1) , coupon_dates[1:], coupon_dates[:-1]))
    auxiliary_=dates[0]
    for i in range(1,len(dates)):
        if control_2[i]==True:
            auxiliary_=ind.iloc[index(ind, dates[i], 0), 3]
        else:
            time.insert(i-1,(1+cupom/100) * ((ind.iloc[index(ind, dates[i], 0), 3] - auxiliary_)/ 252)-1)
    return [0]+(principal * time[-len(dates)+1:]).tolist()


def cdi_plus_coupon(amortization, outstanding, control_2, spread, indexador, dates, coupon_dates, cdi):
    coupon = [0]
    ind = indexador
    principal = (amortization[1:] - outstanding[1:] * control_2[1:])
    spread_interest = list(map(lambda x, y: ((spread + 100) / 100) ** ((ind.iloc[index(ind, x, 0), 3]
                                                                                   - ind.iloc[
                                                                                       index(ind, y, 0), 3]) / 252),
                                          coupon_dates[1:], coupon_dates[:-1]))
    auxiliary_=dates[0]
    for i in range(1,len(dates)):
        if control_2[i]==True:
            auxiliary_=ind.iloc[index(ind, dates[i], 0), 3]
        else:
            spread_interest.insert(i-1,((spread + 100) / 100)**((ind.iloc[index(ind, dates[i], 0), 3]
                                                                                   -auxiliary_)/252))
    spread_interest=np.asarray(spread_interest)
    adjusted_cdi = np.asarray(list(map(lambda x,y: (x*y) - 1, cdi[1:],spread_interest)))
    coupon.extend(principal * np.asarray(adjusted_cdi))
    return coupon

def cdi_percent_coupon(amortization, outstanding, control_2, spread, indexador, dates, cdi):
    coupon = [0]
    ind = indexador
    principal = (amortization[1:] - outstanding[1:] * control_2[1:])
    daily_cdi = np.asarray(list(map(lambda x, y, z: (x ** (1 / (ind.iloc[index(ind, y, 0), 3]
                            - ind.iloc[index(ind, z, 0), 3])) - 1), cdi[1:], dates[1:], dates[:-1])))
    
    adjusted_cdi = np.asarray(list(map(lambda x, y, z: (x * spread / 100 + 1) 
            ** ((ind.iloc[index(ind, y, 0), 3]- ind.iloc[index(ind, z, 0), 3]))
            - 1,daily_cdi, dates[1:], dates[:-1])))
    coupon.extend(principal * adjusted_cdi)
    return coupon

def libor_6_coupon(amortization, outstanding, control_2, libor, spread, freq, dates, coupon_dates):
    principal = (amortization[1:] - outstanding[1:] * control_2[-len(amortization)+1:])
    interest = (((libor*100 + spread)) / 100)
    interest_2 = list(map(lambda x, y, z: x * ((y - z).days / 360 ), interest, coupon_dates[1:], coupon_dates[:-1]))
    interest_2=interest_2[-len(dates)+len(control_2)-sum(control_2):]
    auxiliary_=dates[0]
    for i in range(1,len(dates)):
        if control_2[i]==True:
            auxiliary_=dates[i]
        else:
            interest_2.insert(i-1,interest_2[i-1]*((dates[i] - auxiliary_).days / 360))
    return [0]+(principal * np.asarray(interest_2[-len(dates)+1:])).tolist()

def coupon_values_libor_6(indexador, amortization, outstanding, coupon_dates, dates, spread, freq, data_base,control_2):
    dates_libor=[]
    for i in coupon_dates[:-1]:
        dates_libor.append(i-relativedelta(days=2))
    libor_ = indexador[indexador.loc[:,"PERÍODO"].isin(dates_libor)].loc[:,'LIBOR 6M']
    coupon_values = libor_6_coupon(amortization, outstanding, control_2, libor_, spread, freq, dates,coupon_dates)
    return coupon_values

def sofr_6_coupon(amortization, outstanding, control_2, sofr, spread, freq, dates, coupon_dates):
    principal = (amortization[1:] - outstanding[1:] * control_2[-len(amortization)+1:])
    interest = (((sofr*100 + spread)) / 100)
    interest_2 = list(map(lambda x, y, z: x * ((y - z).days / 360 ), interest, coupon_dates[1:], coupon_dates[:-1]))
    interest_2=interest_2[-len(dates)+len(control_2)-sum(control_2):]
    auxiliary_=dates[0]
    for i in range(1,len(dates)):
        if control_2[i]==True:
            auxiliary_=dates[i]
        else:
            interest_2.insert(i-1,interest_2[i-1]*((dates[i] - auxiliary_).days / 360))
    return [0]+(principal * np.asarray(interest_2[-len(dates)+1:])).tolist()

def coupon_values_sofr_6(indexador, amortization, outstanding, coupon_dates, dates, spread, freq, data_base,control_2):
    dates_sofr=[]
    for i in coupon_dates[:-1]:
        dates_sofr.append(i-relativedelta(days=2))
    sofr_ = indexador[indexador.loc[:,"PERÍODO"].isin(dates_sofr)].loc[:,'SOFR 6M.1']
    coupon_values = sofr_6_coupon(amortization, outstanding, control_2, sofr_, spread, freq, dates,coupon_dates)
    return coupon_values

def libor_3_coupon(amortization, outstanding, control_2, libor, spread, freq, dates, coupon_dates):
    principal = (amortization[1:] - outstanding[1:] * control_2[-len(amortization)+1:])
    interest = (((libor*100 + spread)) / 100)
    interest_2 = list(map(lambda x, y, z: x * ((y - z).days / 360 ), interest, coupon_dates[1:], coupon_dates[:-1]))
    interest_2=interest_2[-len(dates)+len(control_2)-sum(control_2):]
    auxiliary_=dates[0]
    for i in range(1,len(dates)):
        if control_2[i]==True:
            auxiliary_=dates[i]
        else:
            interest_2.insert(i-1,interest_2[i-1]*((dates[i] - auxiliary_).days / 360))
    return [0]+(principal * np.asarray(interest_2[-len(dates)+1:])).tolist()

def coupon_values_libor_3(indexador, amortization, outstanding, coupon_dates, dates, spread, freq, data_base,control_2):
    dates_libor=[]
    for i in coupon_dates[:-1]:
        dates_libor.append(i-relativedelta(days=2))
    libor_ = indexador[indexador.loc[:,"PERÍODO"].isin(dates_libor)].loc[:,'LIBOR 3M']
    coupon_values = libor_3_coupon(amortization, outstanding, control_2, libor_, spread, freq, dates,coupon_dates)
    return coupon_values

def sofr_3_coupon(amortization, outstanding, control_2, sofr, spread, freq, dates, coupon_dates):
    principal = (amortization[1:] - outstanding[1:] * control_2[-len(amortization)+1:])
    interest = (((sofr*100 + spread)) / 100)
    interest_2 = list(map(lambda x, y, z: x * ((y - z).days / 360 ), interest, coupon_dates[1:], coupon_dates[:-1]))
    interest_2=interest_2[-len(dates)+len(control_2)-sum(control_2):]
    auxiliary_=dates[0]
    for i in range(1,len(dates)):
        if control_2[i]==True:
            auxiliary_=dates[i]
        else:
            interest_2.insert(i-1,interest_2[i-1]*((dates[i] - auxiliary_).days / 360))
    return [0]+(principal * np.asarray(interest_2[-len(dates)+1:])).tolist()

def coupon_values_sofr_3(indexador, amortization, outstanding, coupon_dates, dates, spread, freq, data_base,control_2):
    dates_sofr=[]
    for i in coupon_dates[:-1]:
        dates_sofr.append(i-relativedelta(days=2))
    sofr_ = indexador[indexador.loc[:,"PERÍODO"].isin(dates_sofr)].loc[:,'SOFR 3M.1']
    coupon_values = sofr_3_coupon(amortization, outstanding, control_2, sofr_, spread, freq, dates,coupon_dates)
    return coupon_values

def ipca_coupon(amortization, outstanding, control_2, indexador, spread, dates,coupon_dates):
    ind = indexador
    principal = (amortization[1:] - outstanding[1:] * control_2[-len(amortization)+1:])
    adjusted_spread = list(map(lambda y, z: (((spread + 100) / 100) ** ((ind[ind.loc[:,'PERÍODO'] == y].iloc[0, 3] -
                                                           ind[ind.loc[:,'PERÍODO'] == z].iloc[0, 3]) / 252)) - 1, coupon_dates[1:],
                 coupon_dates[:-1]))
    auxiliary_,control_3_1, control_3_2=dates[0],0,0
    for i in range(1,len(dates)):
        while True:
            if coupon_dates[control_3_1]<dates[i]:
                control_3_1+=1
            else:
                break      
        if control_2[i]==True:
            auxiliary_=dates[i]
            control_3_1+=1
        else:
           adjusted_spread.insert(control_3_1+control_3_2-1,((spread + 100) / 100)**((ind[ind.loc[:,'PERÍODO'] == dates[i]].iloc[0, 3] -
                                                           ind[ind.loc[:,'PERÍODO'] == auxiliary_].iloc[0, 3]) / 252)-1)
           control_3_2+=1
    return [0]+(principal * np.asarray(adjusted_spread[-len(dates)+1:])).tolist()

def tjlp_coupon(amortization,outstanding,control_2,spread,coupon_dates,dates, indexador):
    tjlp=list(map(lambda x: 0.06 if x>6 else x/100,indexador[indexador.loc[:,"PERÍODO"].isin(coupon_dates)].loc[:,'TJLP(%).1']))
    principal = (amortization[1:] - outstanding[1:] * control_2[-len(amortization)+1:])
    interest=list(map(lambda x,y,z:((1+x)*1.01*(1+spread/100))**((y-z).days/360)-1,tjlp,coupon_dates[1:],coupon_dates[:-1]))
    auxiliary_,control_3_1, control_3_2=[dates[0], tjlp[0]],0,0
    for i in range(1,len(dates)):
        while True:
            if coupon_dates[control_3_1]<dates[i]:
                control_3_1+=1
            else:
                break      
        if control_2[i]==True:
            auxiliary_=[dates[i],tjlp[control_3_1]]
            control_3_1+=1
        else:
            interest.insert(control_3_1+control_3_2-1, ((1+auxiliary_[1])*1.01*(1+spread/100))**((dates[i]-auxiliary_[0]).days/360)-1)
            control_3_2+=1  
    return [0]+(principal * np.asarray(interest[-len(dates)+1:])).tolist()


def coupon_values_cdi(tipo_de_indexação, amortization, outstanding, coupon_dates, dates, spread, data_base,
                      indexador, control_2):
    cdi_acum = list(indexador[indexador["PERÍODO"].isin(coupon_dates)].iloc[:, 2])
    cdi = [1] + list(map(lambda x, y: x / y, cdi_acum[1:], cdi_acum[:-1]))
    auxiliary_,control_3=[],0
    cdi=cdi[-len(dates)+len(control_2)-sum(control_2)-1:]
    for i in range(1,len(dates)):
        if control_2[i]==True:
            auxiliary_=indexador.iloc[index(indexador, dates[i], 0), 2]
        else:
            cdi.insert(i-control_3,indexador.iloc[index(indexador, dates[i], 0), 2]/auxiliary_)
            control_3+=1
                
    if tipo_de_indexação == "%cdi":
        coupon_values = cdi_percent_coupon(amortization, outstanding, control_2, spread, indexador, dates, cdi)
    else:
        coupon_values = cdi_plus_coupon(amortization, outstanding, control_2, spread, indexador, dates,coupon_dates, cdi)
    return coupon_values

def mapa4_coupon(cupom,freq,outstanding, amortization, control_2):
    interest=(1+cupom/100)**(freq/12)-1
    return[0]+ list(map(lambda x,y,z: (x-y*z)*interest,amortization[1:],outstanding[1:],control_2[1:]))
    

def _interest_(type_, amortization, outstanding, emissão, coupon_dates, dates, freq, data_base, cupom,
             tipo_de_indexação, indexador, nome):
    control_2 = np.asarray(list(map(lambda x: x in coupon_dates, dates)))
    if type_ == "bond":
        result = [bond_coupon(amortization, outstanding, control_2, cupom, freq, dates,coupon_dates)
                  ,outstanding, amortization]
    elif type_=="fixo_360":
        result=[fixo_360_coupon(amortization, outstanding, control_2, cupom,  dates,coupon_dates)
                ,outstanding, amortization]
    elif type_=="fixo_365":
        result=[fixo_365_coupon(amortization, outstanding, control_2, cupom,  dates,coupon_dates)
                ,outstanding, amortization]
    elif type_=="fixo_util":
        result=[fixo_util_coupon(amortization, outstanding, control_2, cupom,  dates,coupon_dates, indexador)
                ,outstanding, amortization]
    elif type_ == "cdi":
        result = [coupon_values_cdi(tipo_de_indexação, amortization, outstanding, coupon_dates, dates, cupom,
                                   data_base, indexador,control_2),outstanding, amortization]
    elif type_ =="libor_6":
        result = [coupon_values_libor_6(indexador, amortization, outstanding, coupon_dates, dates, cupom, freq,
                                     data_base,control_2),outstanding, amortization]
    elif type_ =="sofr_6":
        result = [coupon_values_sofr_6(indexador, amortization, outstanding, coupon_dates, dates, cupom, freq,
                                     data_base,control_2),outstanding, amortization]      
    elif type_ =="libor_3":
        result = [coupon_values_libor_3(indexador, amortization, outstanding, coupon_dates, dates, cupom, freq,
                                     data_base,control_2),outstanding, amortization]
    elif type_ =="sofr_3":
        result = [coupon_values_sofr_3(indexador, amortization, outstanding, coupon_dates, dates, cupom, freq,
                                     data_base,control_2),outstanding, amortization]
    elif type_ == "ipca":
        indexation=ipca_indexation(indexador, outstanding,amortization, dates, data_base)
        outstanding, amortization = indexation[0], indexation[1]
        #workbook=load_workbook(r'C:\Users\u3xp\OneDrive - PETROBRAS\Documents\ipca.xlsx')
        #with pd.ExcelWriter(r'C:\Users\u3xp\OneDrive - PETROBRAS\Documents\ipca.xlsx', engine = 'openpyxl') as writer:
            #writer.book=workbook      
            #indexation[3].to_excel(writer, sheet_name=nome)
        result = [ipca_coupon(amortization, outstanding, control_2, indexador, cupom, dates,coupon_dates)
                  ,indexation[0],indexation[1],indexation[2]]
    elif type_=='tjlp':
        indexation=tjlp_indexation(indexador,coupon_dates,dates,outstanding, amortization, data_base)
        outstanding, amortization = indexation[0], indexation[1]
        result = [tjlp_coupon(amortization,outstanding,control_2,cupom,coupon_dates,dates,indexador),indexation[0],indexation[1]]
    elif type_=='mapa4':
        indexation=mapa4_indexation(cupom, freq, coupon_dates, dates, outstanding, amortization)
        outstanding, amortization = indexation[0], indexation[1]
        result=[mapa4_coupon(cupom,freq,outstanding,amortization,control_2), indexation[0],indexation[1]]
    return result

# FX defs
def swap(indicadores, cash_flow, type_,coupon_dates,data_base,dates,freq,nome, strict_cf):
    from openpyxl import load_workbook
    control_2 = np.asarray(list(map(lambda x: x in coupon_dates, cash_flow.iloc[:,0])))
    col_1 = list(indicadores.columns)
    cdi_acum=list(indicadores[indicadores.loc[:,"PERÍODO"].isin(cash_flow.iloc[:,0])].iloc[:, 2])
    suport_date = indicadores[indicadores.loc[:,col_1[1]] == 1].iloc[0, 0]
    cambio = indicadores[indicadores.loc[:,col_1[1]] == 1].iloc[0, 4]
    col_2 = list(cash_flow.columns)
    flow_2 = cash_flow[cash_flow.loc[:,col_2[0]] >= suport_date].iloc[:]
    control_2=control_2[-len(flow_2):]
    flow_2.iloc[0, -1] = -flow_2.iloc[0, 2] + flow_2.iloc[0, 3]
    flow_3=cash_flow[cash_flow.loc[:,col_2[0]] >= suport_date].iloc[:]
    if type_ in ["ipca","tjlp","mapa4"]:
        deindexer=[1]
        deindexer.extend(list(map(lambda x,y:x/y,flow_2.iloc[1:, 3]-flow_2.iloc[1:, 2],flow_2.iloc[:-1, 3] )))
        for i in range(1,len(deindexer)):
            deindexer[i]=deindexer[i]*deindexer[i-1]
        _ipca_=deindexer
        flow_2.iloc[:, 2]=flow_2.iloc[:, 2]/deindexer
        flow_2.iloc[:, 3] = flow_2.iloc[:, 3] / deindexer
        deindexer=[0]+list(map(lambda x,y:y-x,deindexer[:-1],deindexer[1:]))
        flow_2.iloc[:,4] = flow_2.iloc[:,4] + (flow_2.iloc[:, 2]-flow_2.iloc[:, 3]*control_2)*deindexer
        flow_2.iloc[:, 5]= flow_2.iloc[:, 4] + flow_2.iloc[:, 2]
    flow_2.iloc[0, 5]= pv(indicadores,data_base,coupon_dates,dates,cash_flow)
    di_2 = list(map(lambda x: x / cdi_acum[0], cdi_acum))[-len(flow_2):]
    npv_BRL=sum(flow_2.iloc[:, 5]/di_2) if strict_cf=='no' else sum(flow_3.iloc[:,5]/di_2)
    dia_cupom_2 = cash_flow.iloc[-len(flow_2):,0]
    cf = flow_2.iloc[:, -4:] / cambio
    cf = cf.reset_index(drop=True)
    df = indicadores[indicadores.loc[:,col_1[0]].isin(dia_cupom_2)].loc[:, 'FATOR DE DESCONTO (Cupom Cambial)']
    df=df/df.iloc[0]
    interest = -6
    control_3_1, control_3_2=0,0
    while True:
        if coupon_dates[control_3_1]<dates[control_3_2]:
            control_3_1+=1
        elif coupon_dates[control_3_1]==dates[control_3_2]:
            control_3_1+=1
            control_3_2+=1
        else:
            coupon_dates.insert(control_3_1,coupon_dates[control_3_1-1])
            control_3_2+=1
        if control_3_1==len(coupon_dates) or control_3_2==len(dates):
            break
            coupon_dates.insert(i,coupon_dates[(i-1)])  
    for i in range(1, 14):
        while True:
            interest += 10 ** (-i + 1)
            coupon=[0]+list(map(lambda x,y,z,w,j:(x-y*z)*((1+interest/100)
                                   **((w-j).days/(360/freq))-1),cf.iloc[1:, 0],cf.iloc[1:, 1],control_2[1:],dates[1:],coupon_dates[-len(dates):-1]))            
            cf.loc[:,col_2[4]] = coupon
            cf.loc[:,col_2[5]] = cf.iloc[:, 0] + cf.iloc[:, 2]
            cf.loc[:,col_2[5]][0] = cf.iloc[0, 1]*((1+interest/100)**((dates[0]-coupon_dates[-len(dates)]).days/(360/freq)))+cf.iloc[0, 0]
            npv_USD = sum(list(map(lambda x, y: x * y, list(cf.iloc[:, 3]), list(df))))
            if (npv_USD * cambio) < npv_BRL:
                interest -= 10 ** (-i + 1)
                break
    print(interest*freq)
    
    result = pd.DataFrame(data={'Data': cash_flow.iloc[:, 0]})
    for i in range(2, len(col_2)):
        locals()[f'{col_2[i]}'] = list(cf.loc[:,col_2[i]])
        result.loc[:,f'{col_2[i]}'] = (locals()[f'{col_2[i]}'])
    result.loc[:,'CDI acum']=cdi_acum
    if type_ in ["ipca","tjlp"]:
        result.loc[:,'IPCA/TJLP']=_ipca_
        result.loc[:,'IPCA/TJLP_2']=list(indicadores[indicadores.loc[:,"PERÍODO"].isin(cash_flow.iloc[:,0])].iloc[:, -1])
    result.loc[:,'DU']=list(indicadores[indicadores.loc[:,"PERÍODO"].isin(cash_flow.iloc[:,0])].iloc[:, 3])
    #teste
    flow_2.loc[:,'fator de desconto']=di_2
    flow_2.loc[:,'caixa descontado']=flow_2.iloc[:, -2] / di_2
    cf.loc[:,'fator de desconto']=list(df)
    cf.loc[:,'caixa descontado']=list(map(lambda x, y: x * y, list(cf.iloc[:, 3]), list(df)))
    a=0
    if a==1:
        try:
            workbook=load_workbook(r'C:\Users\u3xp\OneDrive - PETROBRAS\Documents\swap.xlsx')
        except:
            flow_2.to_excel(r'C:\Users\u3xp\OneDrive - PETROBRAS\Documents\swap.xlsx')
            workbook=load_workbook(r'C:\Users\u3xp\OneDrive - PETROBRAS\Documents\swap.xlsx')
        with pd.ExcelWriter(r'C:\Users\u3xp\OneDrive - PETROBRAS\Documents\swap.xlsx', engine = 'openpyxl') as writer:
            for i,j in zip([flow_2, cf],[0,8]):
                writer.book=workbook
                i.to_excel(writer,startcol=j, index=False,sheet_name=nome)    
    #fim do teste
    return [result, ((1+interest/100)**(freq/360)-1)*36000]

def fx(moeda, indexador, cash_flow, data_base,coupon_dates, dates):
    if moeda in ['EUR', 'GBP', 'JPY']:
        fx_ = list(indexador[indexador.loc[:,"PERÍODO"].isin(cash_flow.iloc[:,0])].loc[:,f'USD/{moeda}'])
    else:
        fx_ = [1] * len(cash_flow)
    result = pd.DataFrame(data={'Data': cash_flow.iloc[:, 0]})
    for i in range(4):
        result.loc[:,f'{list(cash_flow.columns)[i+2]}']=cash_flow.iloc[:,i+2]* fx_
    result.iloc[0,-1]=pv(indexador,data_base,coupon_dates,dates,cash_flow)* fx_[0]
    return result

# structural defs
def dt_(string):
    if "/" in list(string)[:4]:
        d1,d2,m1,m2,y1,y2=0,2,3,5,6,10
    else:
        d1,d2,m1,m2,y1,y2=8,10,5,7,0,4
    return datetime.strptime(f'{string[d1:d2]}/{string[m1:m2]}/{string[y1:y2]}', '%d/%m/%Y')

def st_dt(input):
    
    if "/" in list(str(input[0])[:4]):
        d1,d2,m1,m2,y1,y2=0,2,3,5,6,10
    else:
        d1,d2,m1,m2,y1,y2=8,10,5,7,0,4
    return list(map(lambda x: datetime.strptime(f'{str(x)[d1:d2]}/{str(x)[m1:m2]}/{str(x)[y1:y2]}', '%d/%m/%Y'), input))

def b_day(indexador):
    bu_date=[indexador.iloc[0,0]]
    max_=max(indexador.loc[:,'DU Total'])
    bu_date.extend(list(map(lambda x,y: indexador[indexador.loc[:,'DU Total']==x].iloc[0,0] if x!=y else indexador[indexador.loc[:,'DU Total']==(x+1 if x+1<=max_ else x)].iloc[0,0],indexador.loc[:,'DU Total'].iloc[1:],indexador.loc[:,'DU Total'].iloc[:-1])))
    return bu_date

def index(ind,value,col):
    return ind[ind[list(ind.columns)[col] ]== value].index[0]

def creating_lists(data_base,maturity, emissão, freq, amortização, valor_emissão,indexador,b_day_):
    dates, coupon_dates,outstanding, amortization = [data_base],[emissão],[float(valor_emissão)],[0]
    col_amor=list(amortização.columns)
    amortização.iloc[:,0]=list(map(lambda x:b_day_[index(indexador,x,0)],amortização.iloc[:,0]))
    for i in range(int((maturity.year - emissão.year) * freq+(maturity.month-emissão.month)*freq/12)):
        b__day_1=b_day_[index(indexador,emissão + relativedelta(months=12 * (i+1) / freq),0)]
        b__day=b_day_[index(indexador,emissão + relativedelta(months=12 * (i) / freq),0)]
        coupon_dates.append(b__day_1)
        if b__day_1>data_base:
            dates.append(b__day_1)
            if len(amortização[amortização[col_amor[0]]==b__day_1])>0:
                for k in amortização[amortização[col_amor[0]]==b__day_1].iloc[:,1]:
                    amortization.append(-float(k))
            else:
                amortization.append(0)
    control_date=0   
    for i in range(len(dates)-1):
         b__day_1=b_day_[index(indexador, dates [i+control_date+1],0)]
         b__day=b_day_[index(indexador,dates[i+control_date],0)]
         if len(amortização[amortização.loc[:,col_amor[0]].between(b__day,b__day_1, inclusive=False)])>0:
             for j in range(len(amortização[amortização.loc[:,col_amor[0]].between(b__day,b__day_1, inclusive=False)])):
                 dates.insert(i+control_date+1,amortização[amortização[col_amor[0]].between(b__day,b__day_1, inclusive=False)].iloc[j,0])
                 amortization.insert(i+control_date+1,- amortização[amortização[col_amor[0]].between(b__day,b__day_1, inclusive=False)].iloc[j,1])
                 control_date+=1
    for i in range(1,len(dates)):
        outstanding.append(outstanding[i-1]+amortization[i])
    amortization[-1]=-outstanding[-1]
    outstanding[-1]  =0     
    return [dates, amortization, outstanding, coupon_dates]

def conclusion_origin_currency(dates, emissão, name, amortization, outstanding, coupon_values, indexador,type_):
    print(len(dates),len(name),len(amortization),len(outstanding),len(coupon_values))
    cash_flow = pd.DataFrame({'Data': dates, 'Nome': name, 'Amortização': amortization,
                              'Saldo': outstanding, 'Juros': coupon_values})
    cash_flow.loc[:,'Fluxo de Caixa'] = cash_flow["Amortização"] + cash_flow['Juros']
    cash_flow.loc[:,'Data']=cash_flow.loc[:,'Data'].dt.date
    cash_flow.iloc[0, -1] = cash_flow.iloc[0, -3]
    return cash_flow

def npv (cash_flow_usd, discount_factor):
    discount_factor= discount_factor[discount_factor.loc[:,"PERÍODO"].isin(cash_flow_usd.iloc[:,0])].iloc[:,1]
    return cash_flow_usd.iloc[:,9] * list(discount_factor)/discount_factor.iloc[0]
    
def conclusion(cash_flow, cash_flow_usd, discount_factor):
    for i, j in zip(['Amortização USD','Saldo USD','Juros USD','Fluxo de Caixa USD'],[1,2,3,4]):
        cash_flow.loc[:,i]=cash_flow_usd.iloc[:, j]
    cash_flow.loc[:,'Fluxo de Caixa USD FD'] =npv(cash_flow,discount_factor) 
    col_cf = cash_flow.columns
    cash_flow = cash_flow.loc[:,
                              [col_cf[0], col_cf[1], col_cf[2], col_cf[6], col_cf[3], col_cf[7], col_cf[4], col_cf[8], col_cf[5], col_cf[9], col_cf[10]]]
    if cash_flow_usd.shape[1]>5:
        for i in range(5,cash_flow_usd.shape[1]):
            cash_flow.loc[:,f'{list(cash_flow_usd.columns)[i]}']=cash_flow_usd.iloc[:,i]
    return cash_flow